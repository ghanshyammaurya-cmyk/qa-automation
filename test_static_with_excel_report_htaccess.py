import re
import pytest
from datetime import datetime
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# =========================
# CONFIG
# =========================
HTACCESS_USERNAME = "admin"
HTACCESS_PASSWORD = "Intel@2025"

REPORT_FILE = f"Static_Report_HTAccess_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

PAGES = [
    {
        "name": "Industrial Community",
        "url": "https://builders.intel.com/communities/industrial",
        "expected_text": ["Industrial Builders", "Overview", "Solutions", "Training"],
    },
    #{
        #"name": "Network Community",
       # "url": "https://builders-qa.onsumaye.com/communities/network",
        #"expected_text": ["Network Builders", "Overview"],
    #},
    {
        "name": "Retail Community",
        "url": "https://builders.intel.com/communities/retail",
        "expected_text": ["Retail Builders", "Overview"],
    },
    {
        "name": "Video and AI Cities Community",
        "url": "https://builders.intel.com/communities/video-ai-cities",
        "expected_text": ["Video", "Overview"],
    },
    {
        "name": "Healthcare Community",
        "url": "https://builders.intel.com/communities/healthcare-life-sciences",
        "expected_text": ["Healthcare", "Overview"],
    },
]

IGNORED_PATTERNS = [
    r"google-analytics\.com",
    r"googlesyndication\.com",
    r"doubleclick\.net",
    r"sharethis\.com",
    r"onetrust\.com",
    r"cookieconsent",
    r"recaptcha",
    r"google\.com/recaptcha",
    r"dogo\.intel\.com",
    r"googletagmanager\.com",
    r"gstatic\.com",
    r"crwdcntrl\.net",
]

SUSPECT_CONTENT_PATTERNS = [
    r"/api/",
    r"/graphql",
    r"/rest/",
    r"/ajax/",
    r"/json",
    r"search",
    r"listing",
    r"filter",
    r"loadmore",
]

# =========================
# EXCEL SETUP
# =========================
wb = Workbook()
ws = wb.active
ws.title = "Static Validation Report"

headers = [
    "Page Name",
    "Requested URL",
    "Actual URL Loaded",
    "HTTP Status",
    "Visible Content Check",
    "Missing Visible Text",
    "Source Content Check",
    "Missing Source Text",
    "API Check",
    "Suspect API Requests",
    "Redirect Check",
    "Final Result",
    "Remarks",
]

ws.append(headers)

header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# =========================
# HELPERS
# =========================
def is_ignored_request(url: str) -> bool:
    url = url.lower()
    return any(re.search(pattern, url) for pattern in IGNORED_PATTERNS)

def is_suspect_content_request(url: str) -> bool:
    url = url.lower()
    return any(re.search(pattern, url) for pattern in SUSPECT_CONTENT_PATTERNS)

def mark_result_style(row_num: int, final_result: str):
    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    result_cell = ws.cell(row=row_num, column=12)

    if final_result == "PASS":
        result_cell.fill = green_fill
    else:
        result_cell.fill = red_fill

def auto_fit_columns():
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 60)

# =========================
# PYTEST FIXTURE
# =========================
@pytest.fixture(scope="function")
def browser_page():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            ignore_https_errors=True,
            http_credentials={
                "username": HTACCESS_USERNAME,
                "password": HTACCESS_PASSWORD,
            },
        )
        page = context.new_page()
        yield page
        browser.close()

# =========================
# TEST
# =========================
@pytest.mark.parametrize("page_data", PAGES, ids=[p["name"] for p in PAGES])
def test_staticized_pages(browser_page, page_data):
    page = browser_page
    suspect_requests = []
    all_xhr_fetch_requests = []

    def capture_request(request):
        if request.resource_type in ["xhr", "fetch"]:
            url = request.url
            all_xhr_fetch_requests.append(url)
            if not is_ignored_request(url) and is_suspect_content_request(url):
                suspect_requests.append(url)

    page.on("request", capture_request)

    response = page.goto(page_data["url"], wait_until="domcontentloaded", timeout=90000)
    page.wait_for_load_state("networkidle", timeout=90000)

    actual_url = page.url
    body_text = page.locator("body").inner_text()
    html_source = page.content()

    missing_visible = [
        text for text in page_data["expected_text"]
        if text.lower() not in body_text.lower()
    ]

    missing_source = [
        text for text in page_data["expected_text"]
        if text.lower() not in html_source.lower()
    ]

    visible_check = "PASS" if not missing_visible else "FAIL"
    source_check = "PASS" if not missing_source else "FAIL"
    api_check = "PASS" if not suspect_requests else "FAIL"
    redirect_check = "PASS" if actual_url.rstrip("/") == page_data["url"].rstrip("/") else "FAIL"

    remarks = []
    if visible_check == "FAIL":
        remarks.append("Expected text not visible on UI")
    if source_check == "FAIL":
        remarks.append("Expected text missing in HTML source")
    if api_check == "FAIL":
        remarks.append("Content/API requests detected")
    if redirect_check == "FAIL":
        remarks.append("Page redirected to different URL")

    final_result = "PASS" if all([
        response is not None,
        response.status == 200,
        visible_check == "PASS",
        source_check == "PASS",
        api_check == "PASS",
        redirect_check == "PASS",
    ]) else "FAIL"

    row = [
        page_data["name"],
        page_data["url"],
        actual_url,
        response.status if response else "No Response",
        visible_check,
        ", ".join(missing_visible) if missing_visible else "",
        source_check,
        ", ".join(missing_source) if missing_source else "",
        api_check,
        "\n".join(suspect_requests) if suspect_requests else "",
        redirect_check,
        final_result,
        "; ".join(remarks) if remarks else "Static/server-rendered as expected",
    ]

    ws.append(row)
    mark_result_style(ws.max_row, final_result)

    print("\n" + "=" * 80)
    print(f"Testing Page: {page_data['name']}")
    print(f"Requested URL: {page_data['url']}")
    print(f"Actual URL Loaded: {actual_url}")
    print(f"HTTP Status: {response.status if response else 'No Response'}")
    print(f"Visible Content Check: {visible_check}")
    print(f"Source Content Check: {source_check}")
    print(f"API Check: {api_check}")
    print(f"Redirect Check: {redirect_check}")
    print(f"FINAL RESULT: {final_result}")

    assert response is not None, f"No response for {page_data['url']}"
    assert response.status == 200, f"Expected status 200, got {response.status} for {page_data['url']}"
    assert not missing_visible, f"Missing visible text on {page_data['name']}: {missing_visible}"
    assert not missing_source, f"Missing source text on {page_data['name']}: {missing_source}"
    assert not suspect_requests, f"Dynamic content/API requests found on {page_data['name']}: {suspect_requests}"
    assert actual_url.rstrip("/") == page_data["url"].rstrip("/"), (
        f"Redirect mismatch on {page_data['name']}. "
        f"Expected: {page_data['url']}, Actual: {actual_url}"
    )

# =========================
# SAVE REPORT
# =========================
def teardown_module(module):
    auto_fit_columns()
    wb.save(REPORT_FILE)
    print(f"\nExcel Report Generated: {REPORT_FILE}")