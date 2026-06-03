

import openpyxl
from selenium import webdriver
from openpyxl import load_workbook


def get_cell_value(row, cell_index):
    cell = row[cell_index]
    return cell.value.strip() if cell.value else None


def main():
    # Load the Excel file
    file_path = "python "C:\\ghanshyam\\Python""
    workbook = load_workbook(file_path)
    sheet = workbook.active  # Get the first sheet

    # Add headers if not present
    if sheet[1][2].value is None:
        sheet[1][2].value = "Validation Result"  # Column D
    if sheet[1][4].value is None:
        sheet[1][4].value = "Invalid Tags"  # Column F

    # Read the approved tags from column E
    approved_tags = set()
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
        for cell in row:
            if cell.value:
                approved_tags.add(cell.value.strip())

    # Iterate through the courses
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        course_name = get_cell_value(row, 1)  # Column B
        tags = get_cell_value(row, 2)  # Column C

        if course_name and tags:
            tag_list = [tag.strip() for tag in tags.split(",")]
            invalid_tags = [tag for tag in tag_list if tag not in approved_tags]

            row[3].value = "Valid" if not invalid_tags else "Invalid"  # Column D
            if invalid_tags:
                row[5].value = ", ".join(invalid_tags)  # Column F
        else:
            row[3].value = "Invalid (Missing Tags)"  # Column D

    # Save the updated file
    output_path = "C:\\ghanshyam\\Python\\courses_with_results.xlsx"
    workbook.save(output_path)
    workbook.close()
    print("Validation report has been saved to 'courses_with_results.xlsx'")


if __name__ == "__main__":
    main()
